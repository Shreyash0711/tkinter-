from python_files.widget_files.ButtonClass import MyButton
from python_files.widget_files.logger import Log
import tkinter as tk
from tkinter import ttk, messagebox, filedialog

import openpyxl
import json
import os
import shutil
import tempfile


class Excel_tab(ttk.Frame):
    def __init__(self, parent, logger: Log):
        super().__init__(parent)
        self.logger = logger
        self.logger.info_logger("Export tab initialized")

        self.workbook = None
        self.excel_file = None

        # Buttons export and download
        MyButton(self, text="⬇️ Export JSON to Excel", command=self.export_json_to_excel).grid(row=0, column=0, padx=5, pady=5)

        self.download_btn = MyButton(self, text="💾 Download Excel", command=self.download_excel)
        self.download_btn.grid(row=0, column=1, padx=5)
        self.download_btn.config(state="disabled")  # disabled

        
        MyButton(self, text="📂 Open Excel File", command=self.open_excel_file).grid(row=5, column=0, padx=5)

        #  Display exported Excel 
        self.result_box = tk.Text(self, width=45, height=6, state="disabled", wrap="none")
        self.result_box.grid(row=1, column=0, columnspan=3, padx=10, pady=10)

    def export_json_to_excel(self):
        filename = os.path.join("data", "db.json")
        if not os.path.exists(filename):
            messagebox.showerror("Error", "db.json not found.")
            return

        with open(filename, "r") as file:
            try:
                data = json.load(file)
            except json.JSONDecodeError:
                messagebox.showerror("Error", "JSON file is invalid")
                return

        if not data:
            messagebox.showinfo("Info", "No data in JSON file")
            return

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Books"

      
        headers = ["id", "name", "author", "price", "year"]
        ws.append([h.capitalize() for h in headers])

        for item in data:
            ws.append([
                item.get("id", ""),
                item.get("name", ""),
                item.get("author", ""),  
                item.get("price", ""),
                item.get("year", "")
            ])

        # Save to temp file
        temp_dir = tempfile.gettempdir()
        filename = os.path.join(temp_dir, "books_exported.xlsx")
        wb.save(filename)
        self.excel_file = filename

        self.result_box.config(state="normal")
        self.result_box.delete(1.0, tk.END)
        for row in ws.iter_rows(values_only=True):
            self.result_box.insert(tk.END, "\t".join(map(str, row)) + "\n")
        self.result_box.config(state="disabled")

        self.download_btn.config(state="normal")

        self.logger.info_logger(f"JSON data exported to {filename}")
        messagebox.showinfo("Success", f"Exported to temporary file")

    def download_excel(self):
        if not self.excel_file or not os.path.exists(self.excel_file):
            messagebox.showerror("Error", "No Excel file to download.")
            return

        save_path = filedialog.asksaveasfilename(
            defaultextension=".xlsx",
            filetypes=[("Excel files", ".xlsx")],
            title="Save Excel File As",
            initialdir=os.getcwd()
        )
        if save_path:
            try:
                shutil.copy(self.excel_file, save_path)
                self.logger.info_logger(f"Excel downloaded to {save_path}")
                messagebox.showinfo("Downloaded", f"File saved to:\n{save_path}")
            except Exception as e:
                self.logger.error_logger(f"Failed to save Excel: {e}")
                messagebox.showerror("Error", f"Failed to save Excel file:\n{e}")

    def open_excel_file(self):
        file_path = filedialog.askopenfilename(
            title="Select Excel File",
            filetypes=[("Excel files", "*.xlsx *.xls")]
        )
        if not file_path:
            return

        try:
            wb = openpyxl.load_workbook(file_path)
            ws = wb.active  
            valid_cols = ["id", "name", "author", "price", "year"]

            self.result_box.config(state="normal")
            self.result_box.delete(1.0, tk.END)

            headers = next(ws.iter_rows(min_row=1, max_row=1, values_only=True))
            headers_clean = [str(h).strip().lower() for h in headers]

            if headers_clean == valid_cols:
                for row in ws.iter_rows(values_only=True):
                    self.result_box.insert(
                        tk.END,
                        "\t".join([str(cell) if cell is not None else "" for cell in row]) + "\n"
                    )
            else:
                messagebox.showerror("Invalid file", "Column names are not valid")

            self.result_box.config(state="disabled")
            self.logger.info_logger(f"Opened and displayed Excel file: {file_path}")
        except Exception as e:
            self.logger.error_logger(f"Failed to open Excel file: {e}")
            messagebox.showerror("Error", f"Failed to open Excel file:\n{e}")
